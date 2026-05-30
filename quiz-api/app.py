from flask import Flask, request, jsonify, send_file
import cv2
import numpy as np
from pyzbar.pyzbar import decode as pyzbar_decode
import pytesseract
import pandas as pd
from datetime import datetime
import os
import re

app = Flask(__name__)

def load_image(file):
    img_bytes = np.frombuffer(file.read(), np.uint8)
    img = cv2.imdecode(img_bytes, cv2.IMREAD_COLOR)
    return img

def _try_opencv_qr(img):
    detector = cv2.QRCodeDetector()
    data, pts, _ = detector.detectAndDecode(img)
    if data:
        return data
    try:
        wechat = cv2.wechat_qrcode_WeChatQRCode()
        texts, _ = wechat.detectAndDecode(img)
        if texts:
            return texts[0]
    except Exception:
        pass
    return None

def _try_pyzbar(img):
    results = pyzbar_decode(img)
    if results:
        return results[0].data.decode("utf-8")
    return None

def _preprocess_variants(img):
    h, w = img.shape[:2]
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    variants = [img, gray]
    for scale in [2, 3]:
        variants.append(cv2.resize(img, (w*scale, h*scale), interpolation=cv2.INTER_CUBIC))
        variants.append(cv2.resize(gray, (w*scale, h*scale), interpolation=cv2.INTER_CUBIC))
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
    enhanced = clahe.apply(gray)
    variants.append(enhanced)
    variants.append(cv2.resize(enhanced, (w*2, h*2), interpolation=cv2.INTER_CUBIC))
    kernel = np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]])
    variants.append(cv2.filter2D(gray, -1, kernel))
    _, otsu = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    variants.append(otsu)
    variants.append(cv2.resize(otsu, (w*2, h*2), interpolation=cv2.INTER_NEAREST))
    blur = cv2.GaussianBlur(gray, (3, 3), 0)
    adapt = cv2.adaptiveThreshold(blur, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                   cv2.THRESH_BINARY, 11, 2)
    variants.append(adapt)
    for frac_h in [0.30, 0.40]:
        for frac_left in [0.50, 0.55, 0.60]:
            crop_c = img[0:int(h*frac_h), int(w*frac_left):]
            crop_g = gray[0:int(h*frac_h), int(w*frac_left):]
            cw, ch = crop_c.shape[1], crop_c.shape[0]
            variants += [
                crop_c,
                cv2.resize(crop_c, (cw*3, ch*3), interpolation=cv2.INTER_CUBIC),
                crop_g,
                cv2.resize(crop_g, (cw*3, ch*3), interpolation=cv2.INTER_CUBIC)
            ]
    return variants

def decode_qr(img):
    for v in _preprocess_variants(img):
        v = v.astype(np.uint8)
        payload = _try_opencv_qr(v) or _try_pyzbar(v)
        if payload:
            return _parse_qr_payload(payload)
    return None

def _parse_qr_payload(payload):
    result = {"raw": payload, "part1": {}, "part2": {}, "title": ""}
    try:
        parts = payload.split("|")
        result["title"] = parts[0].strip()
        for part in parts[1:]:
            part = part.strip()
            if re.match(r"(?i)part-?i:", part) and not re.match(r"(?i)part-?ii:", part):
                prefix = "part1"
                body = re.sub(r"(?i)part-?i:", "", part).strip()
            elif re.match(r"(?i)part-?ii:", part):
                prefix = "part2"
                body = re.sub(r"(?i)part-?ii:", "", part).strip()
            else:
                continue
            for token in body.split():
                m = re.match(r"Q0*(\d+)=([A-Da-d])", token)
                if m:
                    result[prefix]["Q" + m.group(1)] = m.group(2).upper()
    except Exception as e:
        result["parse_error"] = str(e)
    return result

def extract_student_info(img):
    try:
        h, w = img.shape[:2]
        header = img[0:int(h * 0.35), :]
        gray = cv2.cvtColor(header, cv2.COLOR_BGR2GRAY)
        results_text = []

        enlarged = cv2.resize(gray, (gray.shape[1]*3, gray.shape[0]*3),
                              interpolation=cv2.INTER_CUBIC)
        blur = cv2.GaussianBlur(enlarged, (3, 3), 0)
        _, thresh1 = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        results_text.append(pytesseract.image_to_string(thresh1, config='--oem 3 --psm 6'))

        blur2 = cv2.GaussianBlur(gray, (3, 3), 0)
        thresh2 = cv2.adaptiveThreshold(blur2, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                        cv2.THRESH_BINARY, 11, 2)
        enlarged2 = cv2.resize(thresh2, (thresh2.shape[1]*3, thresh2.shape[0]*3),
                               interpolation=cv2.INTER_CUBIC)
        results_text.append(pytesseract.image_to_string(enlarged2, config='--oem 3 --psm 6'))

        right_half = gray[:, w//2:]
        enlarged3 = cv2.resize(right_half, (right_half.shape[1]*3, right_half.shape[0]*3),
                               interpolation=cv2.INTER_CUBIC)
        _, thresh3 = cv2.threshold(enlarged3, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        results_text.append(pytesseract.image_to_string(thresh3, config='--oem 3 --psm 6'))

        name = ""
        reg_no = ""

        for full_text in results_text:
            lines = [l.strip() for l in full_text.split('\n') if l.strip()]
            for i, line in enumerate(lines):
                ll = line.lower()
                if 'name' in ll and not name:
                    if ':' in line:
                        c = line.split(':', 1)[1].strip()
                        if len(c) > 1:
                            name = c
                    elif i + 1 < len(lines):
                        name = lines[i + 1]
                if not reg_no:
                    if ('reg' in ll or 'registration' in ll) and '#' in line:
                        reg_no = line.split('#', 1)[1].strip()
                    elif ('reg' in ll or 'registration' in ll) and ':' in line:
                        c = line.split(':', 1)[1].strip()
                        if len(c) > 1:
                            reg_no = c
                    elif re.search(r'[A-Z]{2,4}[-_]?[A-Z]{0,3}[-_]?\d{2,4}[-_]\d{3}',
                                   line, re.IGNORECASE):
                        match = re.search(
                            r'[A-Z]{2,4}[-_]?[A-Z]{0,3}[-_]?\d{2,4}[-_]\d{3}',
                            line, re.IGNORECASE)
                        if match:
                            reg_no = match.group(0).upper()
                    elif ('reg' in ll or 'registration' in ll) and i + 1 < len(lines):
                        reg_no = lines[i + 1]
            if name and reg_no:
                break

        if reg_no:
            reg_no = re.sub(r'[^A-Za-z0-9\-]', '', reg_no).upper()

        return {
            "name": name if name else "Not detected",
            "reg_no": reg_no if reg_no else "Not detected"
        }
    except Exception as e:
        return {"name": "OCR error", "reg_no": str(e)}

def read_bubbles(img):
    h, w = img.shape[:2]
    grid = img[int(h * 0.35):int(h * 0.99), int(w * 0.02):int(w * 0.98)]
    gh, gw = grid.shape[:2]

    gray_g = cv2.cvtColor(grid, cv2.COLOR_BGR2GRAY)
    blurred = cv2.GaussianBlur(gray_g, (5, 5), 0)
    edged = cv2.Canny(blurred, 30, 120)
    contours, _ = cv2.findContours(edged, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    best_rect = None
    best_area = 0
    for c in contours:
        peri = cv2.arcLength(c, True)
        approx = cv2.approxPolyDP(c, 0.02 * peri, True)
        area = cv2.contourArea(c)
        if len(approx) == 4 and area > best_area and area > (gh * gw * 0.15):
            best_area = area
            best_rect = approx

    if best_rect is not None:
        pts = best_rect.reshape(4, 2).astype(np.float32)
        rect = np.zeros((4, 2), dtype=np.float32)
        s = pts.sum(axis=1)
        rect[0] = pts[np.argmin(s)]
        rect[2] = pts[np.argmax(s)]
        diff = np.diff(pts, axis=1)
        rect[1] = pts[np.argmin(diff)]
        rect[3] = pts[np.argmax(diff)]
        dst_w, dst_h = 700, 600
        dst = np.array([[0, 0], [dst_w, 0], [dst_w, dst_h], [0, dst_h]], dtype=np.float32)
        M = cv2.getPerspectiveTransform(rect, dst)
        warped = cv2.warpPerspective(grid, M, (dst_w, dst_h))
    else:
        warped = grid

    wh, ww = warped.shape[:2]
    wgray = cv2.cvtColor(warped, cv2.COLOR_BGR2GRAY)
    wblur = cv2.GaussianBlur(wgray, (5, 5), 0)
    _, thresh = cv2.threshold(wblur, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    bubbles = []
    min_r = min(wh, ww) * 0.025
    max_r = min(wh, ww) * 0.13

    for c in contours:
        x, y, bw, bh = cv2.boundingRect(c)
        area = cv2.contourArea(c)
        ratio = bw / float(bh) if bh > 0 else 0
        if (0.55 <= ratio <= 1.8 and min_r <= bw <= max_r
                and min_r <= bh <= max_r and area >= 60):
            cx = x + bw // 2
            cy = y + bh // 2
            roi = thresh[y:y+bh, x:x+bw]
            fill = cv2.countNonZero(roi) / float(bw * bh)
            bubbles.append((cx, cy, bw, bh, fill))

    if len(bubbles) < 8:
        return _grid_fallback_full(thresh, wh, ww)

    bubbles_sorted = sorted(bubbles, key=lambda b: b[1])
    rows = []
    current_row = [bubbles_sorted[0]]
    for b in bubbles_sorted[1:]:
        if b[1] - current_row[-1][1] > wh * 0.05:
            rows.append(current_row)
            current_row = [b]
        else:
            current_row.append(b)
    rows.append(current_row)

    rows = [r for r in rows if len(r) >= 4]
    rows = rows[:8]

    part1 = {}
    part2 = {}

    for ri, row in enumerate(rows):
        qk = "Q" + str(ri + 1)
        row_sorted = sorted(row, key=lambda b: b[0])
        # ── FIXED: use 45%/55% gap instead of 50% to avoid centre Q-number column ──
        left  = [b for b in row_sorted if b[0] < ww * 0.45]
        right = [b for b in row_sorted if b[0] > ww * 0.55]
        part1[qk] = _pick_answer(left)
        part2[qk] = _pick_answer(right)

    return {"part1": part1, "part2": part2}

def _pick_answer(bubbles_in_row):
    options = ["A", "B", "C", "D"]
    if not bubbles_in_row:
        return None
    row = sorted(bubbles_in_row, key=lambda b: b[0])[:4]
    best_idx = None
    best_fill = 0.0
    sec_fill = 0.0
    for i, b in enumerate(row):
        fill = b[4]
        if fill > best_fill:
            sec_fill = best_fill
            best_fill = fill
            best_idx = i
        elif fill > sec_fill:
            sec_fill = fill
    if best_fill < 0.28:
        return None
    if best_fill - sec_fill < 0.10 and sec_fill > 0.18:
        return "INVALID"
    return options[best_idx] if best_idx is not None and best_idx < 4 else None

def _grid_fallback_full(thresh, gh, gw):
    options = ["A", "B", "C", "D"]
    part1 = {}
    part2 = {}
    row_h = gh / 8
    p1_w = gw * 0.42
    p2_start = gw * 0.58
    col_w1 = p1_w / 4
    col_w2 = (gw - p2_start) / 4
    for row in range(8):
        qk = "Q" + str(row + 1)
        y1, y2 = int(row * row_h), int((row + 1) * row_h)
        best_f1, best1 = 0, None
        for col in range(4):
            x1 = int(col * col_w1)
            x2 = int((col + 1) * col_w1)
            cell = thresh[y1:y2, x1:x2]
            f = cv2.countNonZero(cell) / float(cell.size) if cell.size > 0 else 0
            if f > best_f1:
                best_f1, best1 = f, col
        part1[qk] = options[best1] if best1 is not None and best_f1 >= 0.25 else None
        best_f2, best2 = 0, None
        for col in range(4):
            x1 = int(p2_start + col * col_w2)
            x2 = int(p2_start + (col + 1) * col_w2)
            cell = thresh[y1:y2, x1:x2]
            f = cv2.countNonZero(cell) / float(cell.size) if cell.size > 0 else 0
            if f > best_f2:
                best_f2, best2 = f, col
        part2[qk] = options[best2] if best2 is not None and best_f2 >= 0.25 else None
    return {"part1": part1, "part2": part2}

def grade(student_answers, answer_key, negative_marking=False):
    correct = incorrect = unattempted = invalid = 0
    breakdown = {}
    for part in ["part1", "part2"]:
        s_part = student_answers.get(part, {})
        a_part = answer_key.get(part, {})
        for q_num in range(1, 9):
            qk = "Q" + str(q_num)
            key = part + "_" + qk
            student_a = s_part.get(qk)
            correct_a = a_part.get(qk)
            if not correct_a:
                continue
            if student_a == "INVALID":
                invalid += 1
                breakdown[key] = "invalid"
            elif not student_a:
                unattempted += 1
                breakdown[key] = "unattempted"
            elif student_a == correct_a:
                correct += 1
                breakdown[key] = "correct"
            else:
                incorrect += 1
                breakdown[key] = "incorrect"
    total = correct + incorrect + unattempted + invalid
    if negative_marking:
        marks = correct - (0.25 * incorrect) - (0.25 * invalid)
    else:
        marks = correct
    pct = round(marks / total * 100, 1) if total > 0 else 0.0
    letter = ("A" if pct >= 90 else "B" if pct >= 80
              else "C" if pct >= 70 else "D" if pct >= 60 else "F")
    return {
        "correct": correct,
        "incorrect": incorrect,
        "unattempted": unattempted,
        "invalid": invalid,
        "score": str(correct) + "/" + str(total),
        "percentage": pct,
        "grade": letter,
        "breakdown": breakdown
    }

@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "API is running"})

@app.route("/scan", methods=["POST"])
def scan():
    if "image" not in request.files:
        return jsonify({"error": "No image sent"}), 400
    img = load_image(request.files["image"])
    answer_key = decode_qr(img)
    if not answer_key:
        student_info = extract_student_info(img)
        return jsonify({"error": "QR code not found", "student": student_info}), 400
    student_info = extract_student_info(img)
    student_answers = read_bubbles(img)
    neg = "negative" in answer_key.get("raw", "").lower()
    grade_result = grade(student_answers, answer_key, neg)
    return jsonify({
        "student": student_info,
        "answer_key": answer_key,
        "student_answers": student_answers,
        "grade": grade_result
    })

@app.route("/batch", methods=["POST"])
def batch():
    if "images" not in request.files:
        return jsonify({"error": "No images sent"}), 400
    files = request.files.getlist("images")
    all_results = []
    errors = []
    for file in files:
        img = load_image(file)
        answer_key = decode_qr(img)
        if not answer_key:
            errors.append({"file": file.filename, "error": "QR not found"})
            continue
        student_info = extract_student_info(img)
        student_answers = read_bubbles(img)
        neg = "negative" in answer_key.get("raw", "").lower()
        grade_result = grade(student_answers, answer_key, neg)

        title = answer_key.get("title", "")
        quiz_set = ""
        if "Set-" in title:
            quiz_set = title.split("Set-")[-1].split()[0]

        # Build row in exact column order required by assignment
        row = {
            "Quiz":    title,
            "Set":     quiz_set,
            "Class":   "BSE-4A",
            "Subject": "Artificial Intelligence",
            "Name":    student_info.get("name", ""),
            "Reg No":  student_info.get("reg_no", ""),
        }
        # Part1_Q01 ... Part1_Q08
        for q in range(1, 9):
            row["Part1_Q" + str(q).zfill(2)] = student_answers.get("part1", {}).get("Q" + str(q), "")
        # Part2_Q01 ... Part2_Q08
        for q in range(1, 9):
            row["Part2_Q" + str(q).zfill(2)] = student_answers.get("part2", {}).get("Q" + str(q), "")
        # Scoring columns
        row["Correct"]     = grade_result["correct"]
        row["Incorrect"]   = grade_result["incorrect"]
        row["Unattempted"] = grade_result["unattempted"]
        row["Total Marks"] = grade_result["correct"]
        row["Percentage"]  = grade_result["percentage"]
        row["Grade"]       = grade_result["grade"]
        all_results.append(row)

    if not all_results:
        return jsonify({"error": "No valid quizzes", "details": errors}), 400

    df = pd.DataFrame(all_results)

    # Single summary row — class average, highest score, lowest score
    summary = {
        "Quiz":        "SUMMARY",
        "Set":         "",
        "Class":       "BSE-4A",
        "Subject":     "Artificial Intelligence",
        "Name":        "Class Avg: " + str(round(df["Percentage"].mean(), 1)) + "%" +
                       "  |  Highest: " + str(df["Total Marks"].max()) +
                       "  |  Lowest: "  + str(df["Total Marks"].min()),
        "Reg No":      "",
        "Correct":     round(df["Correct"].mean(), 1),
        "Incorrect":   round(df["Incorrect"].mean(), 1),
        "Unattempted": round(df["Unattempted"].mean(), 1),
        "Total Marks": round(df["Total Marks"].mean(), 1),
        "Percentage":  round(df["Percentage"].mean(), 1),
        "Grade":       "",
    }
    for q in range(1, 9):
        summary["Part1_Q" + str(q).zfill(2)] = ""
        summary["Part2_Q" + str(q).zfill(2)] = ""

    df = pd.concat([df, pd.DataFrame([summary])], ignore_index=True)

    # Auto-name file with quiz title + timestamp (as required)
    quiz_title = all_results[0].get("Quiz", "Quiz").replace(" ", "_").replace("|", "")
    os.makedirs("output", exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fn = quiz_title + "_" + ts + ".xlsx"
    filepath = "output/" + fn

    # Style the Excel output
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Results")
        wb = writer.book
        ws = writer.sheets["Results"]
        from openpyxl.styles import PatternFill, Font, Alignment
        # Header row — bold + blue background
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        # Summary rows — yellow background
        summary_fill = PatternFill("solid", fgColor="FFD966")
        for row_idx in range(len(all_results) + 2, len(df) + 2):
            for cell in ws[row_idx]:
                cell.fill = summary_fill
                cell.font = Font(bold=True)
        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 20)
    return jsonify({
        "message": "Processed " + str(len(all_results)) + " quizzes",
        "results": all_results,
        "errors": errors,
        "file": fn
    })

@app.route("/download/<filename>", methods=["GET"])
def download(filename):
    path = "output/" + filename
    if not os.path.exists(path):
        return jsonify({"error": "File not found"}), 404
    return send_file(path, as_attachment=True)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)